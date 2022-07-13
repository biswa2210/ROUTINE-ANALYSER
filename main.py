import  json
# CREATED BY BISWARUP BHATTACHARJEE
facluties=[]
from DataSearching import  Searching
with open('faculties.json', 'r') as json_file:
	json_load = json.load(json_file)
for key in json_load.keys():
    facluties.append(key)
from ttkwidgets.autocomplete import AutocompleteCombobox
from tkinter import *
from tkinter import messagebox
from tkinter_custom_button import TkinterCustomButton
from openpyxl import load_workbook
workbook = load_workbook(filename="DemoExcel.xlsx")
ws = Tk()
ws.iconbitmap("uem.ico")
ws.title('Class Time Table Analyzer')
ws.geometry('500x400')
ws.maxsize(500,350)
ws.minsize(500,350)
ws.config(bg='#DFE7F2')
canvas = Canvas(ws, width = 97, height = 75)
canvas.pack()
img = PhotoImage(file="uem.png")
canvas.create_image(0,0, anchor=NW, image=img)
canvas.place(x=200,y=20)
frame = Frame(ws, bg='#DFE7F2')
frame.pack(expand=True)

Label(
    frame,
    bg='#DFE7F2',
    font = ('Times',21),
    text='Select Faculty\'s Id :-'
    ).pack()

entry = AutocompleteCombobox(
    frame,
    width=30,
    font=('Times', 18),
    completevalues=facluties
    )
entry.pack()
Label(
    frame,
    bg='#DFE7F2',
    font = ('Times',21),
    text='Select Day :- '
    ).pack()
entry2 = AutocompleteCombobox(
    frame,
    width=20,
    font=('Times', 15),
    completevalues=workbook.sheetnames
    )
entry2.pack()

def checkcmbo():
    for fnm in facluties:
        if(fnm==entry.get()):
            f_id=json_load[fnm]
            Searching(f_id,entry2.get())


btn = TkinterCustomButton(text="Fetch Details",corner_radius=15,command=checkcmbo)
btn.place(relx="0.38",rely="0.8")
ws.mainloop()
